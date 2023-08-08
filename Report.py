from tkinter import *

import psycopg2
import datetime as time
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, RIDGE, messagebox

from openpyxl.packaging import workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

def reports():
    current_date_time=time.datetime.now()

    formatted_date_time = current_date_time.strftime("%d/%m/%Y")

    window = tk.Tk()
    window.geometry("600x400")
    window.title("Class Report")


    def day():

        if entry1.get()=="":
            messagebox.showerror("required date","Missing Date")
        elif text.get()=="":
            messagebox.showerror("required class","select class")
        elif text1.get()=="":
            messagebox.showerror("required course","Enter course")
        else:

            date=entry1.get()
            the_class=text.get()
            course=text1.get()

            conn = psycopg2.connect(host="localhost", dbname="apam", port="5432", user="postgres",
                                    password="Zingaro1")
            cur = conn.cursor()
            cur.execute("""SELECT * FROM students where date=""" + "\'" + date + "\'" + """ and class=""""\'" + the_class + "\'" + """ and course=""""\'" + course + "\'")
            r = cur.fetchall()
            print(r)
            daywindow = tk.Tk()
            daywindow.geometry("800x400")
            daywindow.title("Class Report")
            Label(daywindow,text="Report for "+date+" class "+the_class+"").pack()
            table = ttk.Treeview(daywindow, columns=("ID", "Name", "Status","gender"), show="headings")
            table.heading("Name", text="Name",anchor=tk.CENTER)
            table.heading("gender", text="gender")
            table.heading("ID", text="ID")
            table.heading("Status", text="Status")
            table.pack(fill="both", expand=True,pady=20)


            for i in r:
                id = i[0]
                name=i[1]
                status=i[2]
                gender=i[5]
                data = (id,name,status,gender)
                table.insert(parent='', index=0, values=data)
            def savedata():
                number_of_boys_present = 0
                number_of_girls_present = 0
                number_of_boys_absent = 0
                number_of_girls_absent = 0
                loc = filedialog.askdirectory()

                conn = psycopg2.connect(host="localhost", dbname="apam", port="5432", user="postgres", password="Zingaro1")
                cur = conn.cursor()
                cur.execute("""SELECT * FROM students where date=""" + "\'" + date + "\'" + """ and class=""""\'" + the_class + "\'" + """ and course=""""\'" + course + "\'")
                r = cur.fetchall()
                for k in r:
                    print(k[2],k[5])
                    if k[2]=="present" and k[5]=="male":
                        number_of_boys_present+=1
                    if k[2]=="present" and k[5]=="female":
                        number_of_girls_present+=1
                    if k[2]=="absent" and k[5]=="male":
                        number_of_boys_absent+=1
                    if k[2]=="absent" and k[5]=="female":
                        number_of_girls_absent+=1
                print(number_of_boys_present)

                wb = Workbook()
                ws = wb.active
                ws.title = 'Data'
                ws.column_dimensions['B'].width=20
                ws.column_dimensions['E'].width = 20
                ws.merge_cells('A1:B1')
                ws.merge_cells('A2:B2')
                ws.merge_cells('A3:B3')
                ws.merge_cells('A4:B4')
                ws['A1'].value = "Number of men present"
                ws['A2'].value = "Number of women present"
                ws['A3'].value = "Number of men absent"
                ws['A4'].value = "Number of women absent"
                ws['C1'].value = number_of_boys_present
                ws['C2'].value = number_of_girls_present
                ws['C3'].value = number_of_boys_absent
                ws['C4'].value = number_of_girls_absent
                ws['A6'].value = "ID"
                ws['B6'].value = "Name"
                ws['C6'].value = "Status"
                ws['D6'].value = "Month"
                ws['E6'].value = "Date"
                ws['F6'].value = "Gender"
                ws['G6'].value = "Year"
                ws['H6'].value = "Class"
                ws['I6'].value = "Course"
                for i in r:
                    print(i[2],i[5])
                    if i[2]=="present" and i[5]=="male":
                        number_of_boys_present+=1
                    elif i[2]=="present" and i[5]=="female":
                        number_of_girls_present+=1
                    elif i[2]=="absent" and i[5]=="male":
                        number_of_boys_absent+=1
                    else:
                        number_of_girls_absent+=1
                    ws.append(i)
                changedate = date.replace("/", "-")

                t=wb.save(f'{loc}/report_for_{changedate}.xlsx')
                messagebox.showinfo("save report","Your Report has been saved in your selected directory")


            table.bind("<<<TreeviewSelect>>>")
            Submit = tk.Button(daywindow, text="Save report", bg="#CB054A", font=("arial", 15, "italic bold"), relief=RIDGE,
                               activebackground="#7B0519", activeforeground="white", bd=5, width=10,command=savedata)
            Submit.pack()
            window.destroy()


    def on_enter(e):
        entry1.configure(state='normal')
        if entry1.get() == "":
            entry1.insert(0, "11/12/2021")
        else:
            entry1.delete(0, 'end')
            entry1.insert(0, "11/12/2021")


    def enter(e):
        entry1.configure(state='normal')
        if entry1.get() == "":
            entry1.insert(0, "june/2021")
        else:
            entry1.delete(0, 'end')
            entry1.insert(0, "june/2021")

    def delete(e):
        entry1.delete(0,'end')

    label1 = tk.Label(window, text="Report for:", font=("arial", 15, "italic bold"))
    label1.place(x=15, y=20)

    var = tk.IntVar()
    radio1 = tk.Radiobutton(window, text="Day", variable=var, value=1, font=("arial", 15, "bold"))
    radio1.bind("<Button-1>", on_enter)

    radio1.place(x=150, y=20)
    radio2 = tk.Radiobutton(window, text="Month", variable=var, value=2, font=("arial", 15, "bold"))
    radio2.place(x=220, y=20)
    radio2.bind("<Button-1>", enter)

    label1 = tk.Label(window, text="Date :", font=("arial", 15, "italic bold"))
    label1.place(x=15, y=80)
    cntdata = tk.StringVar()
    entry1 = tk.Entry(window, textvariable=cntdata, font=("arial", 20, "italic bold"), bd=2, width=10,state=DISABLED)
    entry1.place(x=100, y=80)
    entry1.bind("<FocusIn>",delete)

    class_options = ['cs1', 'cs2', 'cs3', 'cs4']

    label1 = tk.Label(window, text="Select class :", font=("arial", 15, "italic bold"))
    label1.place(x=15, y=150)

    text = tk.StringVar()
    text.set(class_options[0])
    drop = tk.OptionMenu(window, text, *class_options)
    drop.pack(expand=True)
    drop.configure(bg="black", fg="white", font=("arial", 15, "italic bold"))
    drop["menu"].config(bg="white")
    drop.place(x=180, y=150)

    course_options = ['csm20', 'csm10', 'cs13', 'cs41']

    label0 = tk.Label(window, text="Select course :", font=("arial", 15, "italic bold"))
    label0.place(x=15, y=200)

    text1 = tk.StringVar()
    text1.set(course_options[0])
    drop1 = tk.OptionMenu(window, text1, *course_options)
    drop1.pack(expand=True)
    drop1.configure(bg="black", fg="white", font=("arial", 15, "italic bold"))
    drop1["menu"].config(bg="white")
    drop1.place(x=180, y=200)

    Submit = tk.Button(window, text="View Report", bg="#CB054A", font=("arial", 15, "italic bold"), activebackground=
    "#7B0519", activeforeground="white", bd=5, width=15, command=day)
    Submit.place(x=20, y=280)

    window.mainloop()

